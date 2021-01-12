import openpyxl as opx
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, Alignment, Border, Side
import datetime

wb = opx.Workbook()

#####################################################################################

# 헤더 행 설정

# 헤더 폰트
Default_FONT = Font( name = '나눔바른펜', size=13, bold=True, vertAlign='baseline')

# 헤더 테두리
Default_BORDER = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )

# 헤더 정렬
Default_Allign = Alignment(horizontal='center', vertical='center')

# 헤더 제목 목록 #변경가능
Default_HEADER =['학  번',
                 '이  름', 
                 '성  격(단  어)',
                 '학  업',
                 '특 이 사 항']

# 반복문으로 30번까지 시트를 생성하고, 헤더 입력.
for i in range(0,30): # 0 ~ 30 까지 반복
    wb.create_sheet(index = i, title= str(i+1)) # 셀 제목은 출석번호로. 순서에 맞게

    for alpabhet in ['A','B','C','D','E']:

        wb[str(i + 1)][alpabhet + '1'] = Default_HEADER[['A', 'B', 'C', 'D', 'E'].index(alpabhet)]
        wb[str(i + 1)].column_dimensions['A'].width = 15
        wb[str(i + 1)].column_dimensions['B'].width = 15
        wb[str(i + 1)].column_dimensions['C'].width = 35
        wb[str(i + 1)].column_dimensions['D'].width = 35
        wb[str(i + 1)].column_dimensions['E'].width = 60

        wb[str(i + 1)].row_dimensions[1].height = 20

        wb[str(i + 1)][alpabhet + '1'].font = Default_FONT
        wb[str(i + 1)][alpabhet + '1'].alignment = Default_Allign
        wb[str(i + 1)][alpabhet + '1'].border = Default_BORDER


#####################################################################################

#학번과 이름 입력하기

#학년 input
class_year = input("학년을 입력하세요 ex)1,2,3   :  ")

#반 input
class_number = input("반 번호를 입력하세요 ex)1, 5, 12   :   ")

#학생 명렬표 불러오기
wb_list = opx.load_workbook('./Students_list/1-12학생명렬표.xlsx')
list_sheet = wb_list.active
print(list_sheet.title)

for i in range(0,30): # 0 ~ 30 까지 반복
    wb[str(i + 1)]['A' + str(2)] = int(class_year)*10000 + int(class_number)*100 + i+1
    print(str(list_sheet['A'+str(i+2)].value))

    #중간에 비는 번호가 있을 때,
    if str(list_sheet['A'+str(i+2)].value) == str(i+1) :
        wb[str(i + 1)]['B' + str(2)] = list_sheet['B' + str(i + 2)].value
    elif str(list_sheet['A'+str(i+2)].value) == str(i+2) :
        wb[str(i + 2)]['B' + str(2)] = list_sheet['B' + str(i + 2)].value
    elif str(list_sheet['A'+str(i+2)].value) == str(i+3) :
        wb[str(i + 3)]['B' + str(2)] = list_sheet['B' + str(i + 2)].value
    elif str(list_sheet['A'+str(i+2)].value) == str(i+4) :
        wb[str(i + 4)]['B' + str(2)] = list_sheet['B' + str(i + 2)].value
    else:
        print('없음')

#####################################################################################

#파일 저장 ( 날짜 입력 : 날짜별로 저장안하면 큰일날 수 있음)

time_made = input("오늘날짜를 입력해주세요 ex) 20200225   :   ")

wb.save('.\Students_list/2020_숙명여고_1학년_12반_{}.xlsx'.format(time_made))
# wb.save('E:/@MAHT/@담임/#생활기록부/2020_숙명여고/2020_숙명여고_1학년_12반.xlsx'.format(time_made))
