import openpyxl as opx


# paragraph_1 = sample.readlines()
#     #.txt 파일을 읽을 때에는 줄이 바뀌면, 다른 행으로 인식한다.
#
# all_sentences=[]
#한글로 된 메모 파일을 읽을 때에는, encoding 파라미터를 'utf-8'로 변경해보자.

#################################################################################

#파일 읽어오기

        # 샘플 문장이 들어있는 파일을 읽어온다.
        # 경로와 파일명에 유의한다.

result_wb = opx.load_workbook('.\exel_files\sample.xlsx')

#################################################################################

#HRsht 변수 지정
HRsht = result_wb['HR']

#HRsht에 있는 문단들 다 불러오기
HRparagraphs = []

for i in range(1,HRsht.max_row + 1):
    HRparagraphs.append(HRsht.cell(row=i, column = 2).value)

# print(HRparagraphs)

#################################################################################

#HRsht에 있는 문단들 다 쪼개서, 문장으로 모으기
HRsentences = []

for i in range(0,len(HRparagraphs)): #인덱스에 반복문을 넣어서, 모든 문장들을 분해할 수 있도록 한다.
    if HRparagraphs[i]:
        HRsentences += HRparagraphs[i].split('.')
    else:
        pass
         #all_senteces.append() 클래스메쏘드를 사용하면, 리스트안에 리스트로 원소가 들어가게됨.
         #그러므로, 리스트 연산자 + 를 이용해서 원소들을 다 모아줌.

HRsentences_set = set(HRsentences)
        #중복된 문장을 걸러내기 위해서, set으로 만들어 준다.

# print(HRsentences,'\n','\n')
#################################################################################

# 연결된 파일에서 키워드를 가져와 문장 찾기.
Student_Memo_Book=opx.load_workbook('./Students_list/2020_숙명여고_1학년_12반.xlsx')

# student_sheet_list = Student_Memo_Book.sheetnames

# 번호를 입력하면, '성 격'열(C열)에 있는 성격들을 대상으로 예시문장을 찾아준다.
def char_sentence(a):
    char_list = []
    char_list_1 = []

    if type(Student_Memo_Book[a]['C2'].value)== str :
            # C2 셀에 내용이 있다면, 그 타입은 문자열일 것이다.
            # C2.value의 타입이 문자열일 때만 다음을 진행.

        for Memo in Student_Memo_Book[a]['C']:
                char_list_1 += str(Memo.value).split('.')
                    #문자열.replace 메쏘드로 띄어쓰기를 없앤다. : ' '를 ''로 바꿔줄 수도 있다.
                        # 이 경우, 띄어쓰기가 들어간 어구로는 검색을 하지 않는다.
                    #문자열.split 메쏘드로 '.'를 기준으로 끊어 문자열의 리스트를 만든다.

    else:
        char_list_1 = ['성  격']
        print('정보없음')

                        # print(char_list_1)

    for Memo_2 in char_list_1:
            # 일차적인 만든 리스트에 있는 문자열을 대상으로 ','기준으로 split하여
            # 최종 char_list 를 만든다.
            char_list.append(Memo_2.replace(' ','').split(','))
            # 공백을 제거한 후 ','를 기준으로 split

    # 학번과 이름을 print 한다.
    print('\n')

    print(Student_Memo_Book[a]['A2'].value,'\t',
          '  '.join(Student_Memo_Book[a]['B2'].value))
    print('\n')


    # char_list에 있는 성격 요소들을 print들을 나열한다.
    for char in char_list:
        if char ==['성격(단어)']:
            print('성  격 : ', end=' ')
        else:
            print(char, end=', ')

    print('\n')


    # char_list 에 '성격(단어)'이 들어있으므로, 제거한다.('성격(단어)'이 들어간 문장을 만들 필요는 없음)
    char_list.remove(['성격(단어)'])


    # char_list에 있는 성격들을 키워드로 하여, 예시문장들을 찾는다.
    for i in range(1, len(char_list)+1):
        key_word_list = char_list[i-1]
                # (예비용) 그냥 단어로 찾는 걸로 바꿀 때 : key_word= char_list[i-1]
        print('-'*15,'<',end='')
        print(key_word_list, end='')
        print('>','단어가 들어간 문장 예시.','-'*15,'\n')

        for sentence in HRsentences_set:
            # Random으로 섞을지 말 지 생각을 해보는 게 좋을 듯.
                # def find_many(sentence, word_list):
                    # sentence는 문자열
                    # word_list는 찾을 문자가 들어있는 리스트.

                discrimination = []

                for i in range(len(key_word_list)):
                    if key_word_list[i] in sentence:
                        discrimination.append('T')
                    else:
                        discrimination.append('F')
                # print(discrimination)

                if 'F' in discrimination:
                    None
                else:
                    print('\t',sentence.strip() + '.''\n')

                # (예비용) 다 뜯어서 하나씩 찾는 걸로 바꿀 때,
                # if sentence.find(key_word)>=0:
                #         print('\t',sentence+'.''\n')
                # if sentence.find(key_word)<0:
                #         None



def study_sentence(a): # 학업에 관련된 문장을 출력하는 함수를 정의한다.
    study_list = []
    study_list_1 = []

    if type(Student_Memo_Book[a]['D2'].value)== str :
            # C2 셀에 내용이 있다면, 그 타입은 문자열일 것이다.
            # C2.value의 타입이 문자열일 때만 다음을 진행.

        for Memo in Student_Memo_Book[a]['D']:
                study_list_1 += str(Memo.value).split('.')
                    #문자열.replace 메쏘드로 띄어쓰기를 없앤다. : ' '를 ''로 바꿔줄 수도 있다.
                        # 이 경우, 띄어쓰기가 들어간 어구로는 검색을 하지 않는다.
                    #문자열.split 메쏘드로 '.'를 기준으로 끊어 문자열의 리스트를 만든다.

    else:
        study_list_1 = ['학  업']
        print('정보없음')

    for Memo_2 in study_list_1:
            # 일차적인 만든 리스트에 있는 문자열을 대상으로 ','기준으로 split하여
            # 최종 study_list 를 만든다.
            study_list.append(Memo_2.replace(' ','').split(','))
            # 공백을 제거한 후 ','를 기준으로 split

    # 학번과 이름을 print 한다.
    print('\n')

    print(Student_Memo_Book[a]['A2'].value,'\t',
          '  '.join(Student_Memo_Book[a]['B2'].value))
    print('\n')

    for study in study_list:
        if study ==['학업']:
            print('학  업 : ', end=' ')
        else:
            print(study, end=', ')

    print('\n')

    # char_list 에 '성격(단어)'이 들어있으므로, 제거한다.('성격(단어)'이 들어간 문장을 만들 필요는 없음)
    study_list.remove(['학업'])


    # char_list에 있는 성격들을 키워드로 하여, 예시문장들을 찾는다.
    for i in range(1, len(study_list)+1):
        key_word_list = study_list[i-1]
                # (예비용) 그냥 단어로 찾는 걸로 바꿀 때 : key_word= char_list[i-1]
        print('-'*15,'<',end='')
        print(key_word_list, end='')
        print('>','단어가 들어간 문장 예시.','-'*15,'\n')

        for sentence in HRsentences_set:
            # Random으로 섞을지 말 지 생각을 해보는 게 좋을 듯.
                # def find_many(sentence, word_list):
                    # sentence는 문자열
                    # word_list는 찾을 문자가 들어있는 리스트.

                discrimination = []

                for i in range(len(key_word_list)):
                    if key_word_list[i] in sentence:
                        discrimination.append('T')
                    else:
                        discrimination.append('F')
                # print(discrimination)

                if 'F' in discrimination:
                    None
                else:
                    print('\t',sentence.strip() + '.''\n')

# 정의된 함수를 사용한다.
# 학생 번호를 입력한다.
while True:
    student_number = input('학생의 번호를 입력해주세요. ex)27  :  ')
    char_sentence(student_number)
    print('*'*50)
    print('*'*50)
    print('*'*50)
    study_sentence(student_number)
    print('<'*15 + 'OVER' + '>'*15)

#################################################################################
    # 마침표과 쉼표를 기준으로 split하지 않은 코딩
                # char_list = Student_Memo_Book[student_number]['C']
                #
                # print(Student_Memo_Book[student_number]['A2'].value,
                #       Student_Memo_Book[student_number]['B2'].value)
                #
                # print('\n')
                #
                #
                #
                # for char in char_list:
                #     if char.value =='성  격':
                #         print('성  격 : ', end=' ')
                #     else:
                #         print(char.value, end=', ')
                #
                # print('\n')
                #
                # for i in range(1,Student_Memo_Book[student_number].max_row):
                #     key_word= str(Student_Memo_Book[student_number]['C'+str(i+1)].value)
                #     print('-'*30,'<'+key_word+'>','단어가 들어간 문장 예시.','-'*30,'\n')
                #
                #     for sentence in HRsentences:
                #         if sentence.find(key_word)>=0:
                #                 print('\t',sentence+'.''\n')
                #         if sentence.find(key_word)<0:
                #                 None
