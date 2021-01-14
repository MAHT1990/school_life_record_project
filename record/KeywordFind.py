import openpyxl as opx


# paragraph_1 = sample.readlines()
#     #.txt 파일을 읽을 때에는 줄이 바뀌면, 다른 행으로 인식한다.
#
# all_sentences=[]
#한글로 된 메모 파일을 읽을 때에는, encoding 파라미터를 'utf-8'로 변경해보자.

#################################################################################

#파일 읽어오기
result_wb = opx.load_workbook('record/exel_files/sample.xlsx')

#################################################################################

#HRsht 변수 지정
HRsht = result_wb['HR']

#HRsht에 있는 문단들 다 불러오기
HRparagraphs = []

for i in range(1,HRsht.max_row + 1):
    HRparagraphs.append(HRsht.cell(row=i, column = 2).value)

# print(HRparagraphs)

#################################################################################

#HRsht에 있는 문단들 다 쪼개서, 문장으로 모으기
HRsentences = []

for i in range(0,len(HRparagraphs)): #인덱스에 반복문을 넣어서, 모든 문장들을 분해할 수 있도록 한다.
     HRsentences += HRparagraphs[i].split('.')
         #all_senteces.append() 클래스메쏘드를 사용하면, 리스트안에 리스트로 원소가 들어가게됨.
         #그러므로, 리스트 연산자를 이용해서 원소들을 다 모아줌.

HRsentences_set = set(HRsentences)

#################################################################################

#키워드를 가진 문장 찾기.

def KeywordFind(kw):
    items=[]
    for sentence in HRsentences_set:
        if sentence.find(kw)>=0:
            items.append(sentence)
        else:
            None
    return items

#
# text용
#
# key_word=input('키워드를 입력하세요......')
# print(KeywordFind(key_word))

# while True:
#     key_word=input('키워드를 입력하세요......')
#
#     for sentence in HRsentences_set:
#         if sentence.find(key_word)>=0:
#             print(sentence+'.','\n')
#         if sentence.find(key_word)<0:
#             None
