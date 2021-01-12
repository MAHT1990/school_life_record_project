import openpyxl as opx

sample = open('E:/@MAHT/@담임/#생활기록부/2020_숙명여고/문장_분석_프로그램.txt',encoding='utf-8')
result_wb = opx.load_workbook('E:\@MAHT\@담임\#생활기록부\샘플문장.xlsx')
#한글로 된 메모 파일을 읽을 때에는, encoding 파라미터를 'utf-8'로 변경해보자.

paragraph_1 = sample.readlines()
    #.txt 파일을 읽을 때에는 줄이 바뀌면, 다른 행으로 인식한다.


paragraph_2=[]

# 인덱스에 반복문을 넣어서, 이름과 내용을 분리시킨다.
for i in range(0,len(paragraph_1)):
    paragraph_2.append(paragraph_1[i].split(':'))

for i in range(1, len(paragraph_1)):
    result_wb['담임'].cell(row=i ,column=1).value = paragraph_2[i - 1][0]
    result_wb['담임'].cell(row=i, column=2).value = paragraph_2[i - 1][1]

sample.close()
result_wb.save('E:\@MAHT\@담임\#생활기록부\샘플문장1.xlsx')

#함수로 정의를 할 때에는, 파일명과 쉬트이름을 변수로 하여 해당 쉬트에 정리가 될 수 있도록.